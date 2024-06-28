"""
1- entrar na planilha e extrair cpf
2- entrar no site com o cpf da planilha para pesquisar o status do pagamento
3- verificar o estado do pagamento
4- se em dia, pegar data e o metodo de pagamento
5- se estiver atrasado, colocar status como pendente
6- inserir essas novas informacoes (nome, cpf, vencimento, status, e caso esteja em dia, data e metodo pagamento(cartao ou boleto)) em nova planilha
7- repetir ate ultimo cliente
"""

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1-
client_sheet = openpyxl.load_workbook('data\dados_clientes.xlsx')
client_page = client_sheet['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(5)
    
for row in client_page.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = row
    # 2-
    search_place = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(1)
    search_place.clear()
    search_place.send_keys(cpf)
    sleep(1)
    # 3-
    search_button = driver.find_element(By.XPATH,"//buttom[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    search_button.click()
    sleep(4)
    # 4-
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if status.text == 'em dia':
        payment_date = driver.find_element(By.XPATH,"//p[@id='paymentDate]")
        payment_method = driver.find_element(By.XPATH,"//p[@id='paymentMethod]")
        payment_date_clear = payment_date.text.split()[3]
        payment_method_clear = payment_method.text.split()[3]
        
        finishing_sheet = openpyxl.load_workbook('data\planilha fechamento.xlsx')
        finishing_page = finishing_sheet['Sheet1']
        
        finishing_page.append([nome, valor, cpf, vencimento, 'em dia', payment_date_clear, payment_method_clear])
        finishing_sheet.save('planilha fechamento.xlsx')
    # 5-
    else:
        finishing_sheet = openpyxl.load_workbook('data\planilha fechamento.xlsx')
        finishing_page = finishing_sheet['Sheet1']
        
        finishing_page.append([nome, valor, cpf, vencimento, 'pendente'])
        finishing_sheet.save('planilha fechamento.xlsx')